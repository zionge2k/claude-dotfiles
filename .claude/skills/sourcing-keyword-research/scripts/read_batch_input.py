#!/usr/bin/env python3
# /// script
# dependencies = ["openpyxl"]
# ///
"""Parse batch input file (md/csv/xlsx) and output JSON.

Usage:
    uv run read_batch_input.py input.md
    uv run read_batch_input.py input.csv
    uv run read_batch_input.py input.xlsx

Output: JSON array of {"set_name", "product_name", "components"} objects.
"""
import csv
import json
import re
import sys
from pathlib import Path


def parse_markdown(text):
    """Parse markdown table with 세트명, (상품명), 구성품 columns."""
    lines = [l.strip() for l in text.strip().splitlines() if l.strip()]
    header_idx = None
    for i, line in enumerate(lines):
        if "|" in line and "세트명" in line:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("마크다운 테이블에서 '세트명' 헤더를 찾을 수 없습니다")

    headers = [h.strip() for h in lines[header_idx].strip("|").split("|")]
    data_start = header_idx + 2

    results = []
    for line in lines[data_start:]:
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        row = dict(zip(headers, cells))
        results.append(_normalize_row(row))
    return results


def parse_csv_file(filepath):
    """Parse CSV with 세트명, (상품명), 구성품 columns."""
    results = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            results.append(_normalize_row(row))
    return results


def parse_xlsx(filepath):
    """Parse XLSX with 세트명, (상품명), 구성품 columns."""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    results = []
    for row in rows[1:]:
        cells = [str(c).strip() if c else "" for c in row]
        row_dict = dict(zip(headers, cells))
        results.append(_normalize_row(row_dict))
    wb.close()
    return results


def _normalize_row(row):
    """Normalize row dict to standard format.

    Supports two formats:
    1. Single column: 세트명, (상품명), 구성품
    2. Multiple columns: 세트명, 구성품1, 구성품2, 구성품3, ...
    """
    set_name = row.get("세트명", "").strip()
    product_name = row.get("상품명", "").strip()

    # Format 1: Single "구성품" column with comma-separated values
    if "구성품" in row:
        components_raw = row.get("구성품", "").strip()
        components = [c.strip() for c in re.split(r"[,，]", components_raw) if c.strip()]
    # Format 2: Multiple "구성품N" columns
    else:
        components = []
        for key, value in row.items():
            if key.startswith("구성품") and value:
                val = str(value).strip()
                if val:
                    components.append(val)

    return {
        "set_name": set_name,
        "product_name": product_name if product_name else set_name,
        "components": components,
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: uv run read_batch_input.py <input_file>", file=sys.stderr)
        sys.exit(1)

    filepath = Path(sys.argv[1])
    if not filepath.exists():
        print(f"Error: {filepath} not found", file=sys.stderr)
        sys.exit(1)

    ext = filepath.suffix.lower()
    if ext == ".md":
        text = filepath.read_text(encoding="utf-8")
        results = parse_markdown(text)
    elif ext == ".csv":
        results = parse_csv_file(filepath)
    elif ext == ".xlsx":
        results = parse_xlsx(filepath)
    else:
        print(f"Error: unsupported format '{ext}' (md/csv/xlsx only)", file=sys.stderr)
        sys.exit(1)

    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
